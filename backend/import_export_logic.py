import pandas as pd

from backend.models import Component
from backend.database import get_session
from backend.component_factory import ComponentFactory
from backend.exceptions import DatabaseError, InvalidInputError, ComponentError

from pandas import ExcelWriter
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill

EXCEL_COLUMNS = ["Part Number", "Type", "Value", "Quantity", "Purchase Link", "Datasheet Link"]
REQUIRED_IMPORT_COLUMNS = ["Part Number", "Type", "Value", "Quantity"]


def export_to_excel(filename: str) -> bool | None:
    components_data = []
    session = get_session()
    try:
        all_components = session.query(Component).order_by(Component.part_number).all()
        for comp in all_components:
            components_data.append({
                "Part Number": comp.part_number,
                "Type": comp.component_type,
                "Value": comp.value,
                "Quantity": comp.quantity,
                "Purchase Link": comp.purchase_link,
                "Datasheet Link": comp.datasheet_link,
            })
    except Exception as e:
        raise DatabaseError(f"Failed to fetch components for export: {e}") from e
    finally:
        session.close()

    df = pd.DataFrame(components_data, columns=EXCEL_COLUMNS)

    try:
        with ExcelWriter(filename, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Inventory', index=False, header=True)
            worksheet = writer.sheets['Inventory']

            for column_cells in worksheet.columns:
                header_text = column_cells[0].value or ""
                max_length = len(str(header_text))
                column_letter = get_column_letter(column_cells[0].column)

                for cell in column_cells:
                    try:
                        cell_text = str(cell.value)
                        if cell_text and len(cell_text) > max_length:
                            max_length = len(cell_text)
                    except (ValueError, TypeError):
                        pass

                adjusted_width = (max_length + 2) * 1.2
                worksheet.column_dimensions[column_letter].width = adjusted_width

            header_font = Font(bold=True, color="FFFFFF")
            header_alignment = Alignment(horizontal='center', vertical='center')
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

            for cell in worksheet[1]:
                cell.font = header_font
                cell.alignment = header_alignment
                cell.fill = header_fill

            worksheet.auto_filter.ref = worksheet.dimensions

        return True

    except IOError as e:
        raise IOError(f"Failed to write Excel file '{filename}': {e}") from e
    except Exception as e:
        raise Exception(f"An unexpected error occurred during Excel export formatting/writing: {e}") from e


def import_from_excel(filename: str) -> bool | None:
    try:
        df = pd.read_excel(filename, engine='openpyxl')
    except FileNotFoundError:
        raise FileNotFoundError(f"Import file not found: {filename}")
    except Exception as e:
        raise InvalidInputError(f"Failed to read or parse Excel file '{filename}': {e}") from e

    missing_cols = [col for col in REQUIRED_IMPORT_COLUMNS if col not in df.columns]
    if missing_cols:
        raise InvalidInputError(f"Import file '{filename}' is missing required columns: {', '.join(missing_cols)}")

    components_to_add = []
    for i, (index, row) in enumerate(df.iterrows()):
        try:
            part_number = str(row["Part Number"]).strip()
            component_type = str(row["Type"]).strip().lower()
            value = str(row["Value"]).strip()
            quantity = int(row["Quantity"])

            purchase_link = row.get("Purchase Link")
            purchase_link = str(purchase_link).strip() if pd.notna(purchase_link) else None

            datasheet_link = row.get("Datasheet Link")
            datasheet_link = str(datasheet_link).strip() if pd.notna(datasheet_link) else None

            if not part_number:
                raise ValueError("Part Number cannot be empty.")
            if not component_type:
                raise ValueError("Type cannot be empty.")
            if not value:
                raise ValueError("Value cannot be empty.")
            if quantity < 0:
                raise ValueError("Quantity cannot be negative.")

            components_to_add.append({
                'part_number': part_number,
                'component_type': component_type,
                'value': value,
                'quantity': quantity,
                'purchase_link': purchase_link if purchase_link else None,
                'datasheet_link': datasheet_link if datasheet_link else None,
            })

        except (KeyError, ValueError, TypeError) as e:
            raise InvalidInputError(f"Invalid data found in row {i + 2} of '{filename}': {e}") from e
        except Exception as e:
            raise Exception(f"Unexpected error processing row {i + 2} of '{filename}': {e}") from e

    session = get_session()
    try:
        _ = session.query(Component).delete()

        for comp_data in components_to_add:
            try:
                component = ComponentFactory.create_component(
                    component_type=comp_data['component_type'],
                    part_number=comp_data['part_number'],
                    value=comp_data['value'],
                    quantity=comp_data['quantity'],
                    purchase_link=comp_data['purchase_link'],
                    datasheet_link=comp_data['datasheet_link']
                )
                session.add(component)
            except ValueError as e:
                session.rollback()
                raise ComponentError(
                    f"Failed to create component for Part Number '{comp_data['part_number']}': {e}") from e

        session.commit()
        return True

    except ComponentError as e:
        raise e
    except Exception as e:
        session.rollback()
        raise DatabaseError(f"Database error during import: {e}") from e
    finally:
        session.close()
